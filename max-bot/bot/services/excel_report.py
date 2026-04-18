"""Генерация Excel-отчётов для администраторов."""

import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_DATE_FMT = "YYYY-MM-DD HH:MM"

MODE_LABELS = {
    "transcribe": "Распознавание",
    "theses": "Тезисы",
    "protocol": "Протокол",
    "custom": "Свой вариант",
}


def _apply_headers(ws, headers: list[str], row: int = 1):
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _fmt_dt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _tmp_path(prefix: str) -> Path:
    f = tempfile.NamedTemporaryFile(prefix=prefix, suffix=".xlsx", delete=False)
    f.close()
    return Path(f.name)


# ── Отчёт 1: Общая статистика ──

def build_overview_report(stats: dict) -> Path:
    """3 листа: Сводка / По дням / Режимы."""
    wb = Workbook()

    # ── Лист 1: Сводка ──
    ws1 = wb.active
    ws1.title = "Сводка"
    _apply_headers(ws1, ["Показатель", "Значение"])
    summary_rows = [
        ("Всего пользователей", stats.get("total_users", 0)),
        ("Активных подписок", stats.get("active_subscriptions", 0)),
        ("Выручка (оплаченные заказы), ₽", stats.get("total_revenue_rub", 0)),
        ("Всего транскрибаций", stats.get("total_transcriptions", 0)),
        ("Всего минут обработано", round(stats.get("total_minutes", 0), 1)),
        ("Дата отчёта", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r, (k, v) in enumerate(summary_rows, 2):
        ws1.cell(row=r, column=1, value=k)
        ws1.cell(row=r, column=2, value=v)
    _autofit(ws1)

    # ── Лист 2: По дням (30 дней) ──
    ws2 = wb.create_sheet("По дням (30 дн)")
    _apply_headers(ws2, ["Дата", "Новых пользователей", "Транскрибаций", "Выручка, ₽"])
    daily = stats.get("daily", [])
    for r, row in enumerate(daily, 2):
        ws2.cell(row=r, column=1, value=str(row.get("date", "")))
        ws2.cell(row=r, column=2, value=row.get("new_users", 0))
        ws2.cell(row=r, column=3, value=row.get("transcriptions", 0))
        ws2.cell(row=r, column=4, value=row.get("revenue", 0))
    _autofit(ws2)

    # ── Лист 3: Режимы ──
    ws3 = wb.create_sheet("Режимы")
    _apply_headers(ws3, ["Режим", "Код", "Количество", "% от всех"])
    modes = stats.get("modes", [])
    total_trans = stats.get("total_transcriptions", 1) or 1
    for r, row in enumerate(modes, 2):
        code = row.get("mode", "")
        count = row.get("count", 0)
        ws3.cell(row=r, column=1, value=MODE_LABELS.get(code, code))
        ws3.cell(row=r, column=2, value=code)
        ws3.cell(row=r, column=3, value=count)
        ws3.cell(row=r, column=4, value=round(count / total_trans * 100, 1))
    _autofit(ws3)

    path = _tmp_path("admin_overview_")
    wb.save(path)
    return path


# ── Отчёт 2: Биллинг пользователя ──

def build_user_billing_report(data: dict) -> Path:
    """3 листа: Профиль / Платежи / Транскрибации."""
    wb = Workbook()
    user = data.get("user") or {}
    stats = data.get("stats") or {}
    sub = data.get("subscription")

    # ── Лист 1: Профиль ──
    ws1 = wb.active
    ws1.title = "Профиль"
    _apply_headers(ws1, ["Поле", "Значение"])
    profile_rows = [
        ("user_id", user.get("user_id", "")),
        ("Username", user.get("username", "")),
        ("Имя", user.get("first_name", "")),
        ("Дата регистрации", _fmt_dt(user.get("created_at"))),
        ("Баланс, ₽", user.get("star_balance", 0)),
        ("Пробный файл использован", "Да" if user.get("trial_used") else "Нет"),
        ("Бесплатные минуты", round(float(user.get("free_minutes") or 0), 1)),
        ("", ""),
        ("— Подписка —", ""),
        ("Активна", "Да" if sub else "Нет"),
        ("Действует до", _fmt_dt(sub.get("expires_at")) if sub else "—"),
        ("Минут всего", sub.get("minutes_total", "—") if sub else "—"),
        ("Минут использовано", round(float(sub.get("minutes_used") or 0), 1) if sub else "—"),
        ("", ""),
        ("— Итого —", ""),
        ("Транскрибаций", stats.get("transcriptions", 0)),
        ("Обработано минут", round(float(stats.get("total_seconds") or 0) / 60, 1)),
        ("Потрачено, ₽", stats.get("total_stars_spent", 0)),
        ("Оплачено, ₽", stats.get("total_stars_paid", 0)),
    ]
    for r, (k, v) in enumerate(profile_rows, 2):
        ws1.cell(row=r, column=1, value=k)
        ws1.cell(row=r, column=2, value=v)
    _autofit(ws1)

    # ── Лист 2: Платежи ──
    ws2 = wb.create_sheet("Платежи")
    _apply_headers(ws2, ["order_id", "Тип", "Сумма, ₽", "Статус", "ID Tinkoff", "Дата"])
    for r, row in enumerate(data.get("orders", []), 2):
        ws2.cell(row=r, column=1, value=row.get("order_id", ""))
        ws2.cell(row=r, column=2, value=row.get("payment_type", ""))
        ws2.cell(row=r, column=3, value=row.get("amount_rub", 0))
        ws2.cell(row=r, column=4, value=row.get("status", ""))
        ws2.cell(row=r, column=5, value=row.get("tinkoff_payment_id", ""))
        ws2.cell(row=r, column=6, value=_fmt_dt(row.get("created_at")))
    _autofit(ws2)

    # ── Лист 3: Транскрибации ──
    ws3 = wb.create_sheet("Транскрибации")
    _apply_headers(ws3, ["Файл", "Режим", "Мин", "Стоимость, ₽", "Пробный", "Дата"])
    for r, row in enumerate(data.get("transcriptions", []), 2):
        mode_code = row.get("mode", "transcribe")
        ws3.cell(row=r, column=1, value=row.get("file_name", ""))
        ws3.cell(row=r, column=2, value=MODE_LABELS.get(mode_code, mode_code))
        ws3.cell(row=r, column=3, value=round(float(row.get("duration_sec") or 0) / 60, 1))
        ws3.cell(row=r, column=4, value=row.get("stars_spent", 0))
        ws3.cell(row=r, column=5, value="Да" if row.get("is_trial") else "Нет")
        ws3.cell(row=r, column=6, value=_fmt_dt(row.get("created_at")))
    _autofit(ws3)

    user_id = user.get("user_id", "unknown")
    path = _tmp_path(f"admin_user_{user_id}_")
    wb.save(path)
    return path


# ── Отчёт 3: Все платежи ──

def build_payments_report(rows: list[dict]) -> Path:
    """1 лист со всеми заказами Tinkoff."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Платежи"
    _apply_headers(ws, ["order_id", "user_id", "Username", "Имя", "Тип", "Сумма, ₽", "Статус", "ID Tinkoff", "Дата"])
    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=row.get("order_id", ""))
        ws.cell(row=r, column=2, value=row.get("user_id", ""))
        ws.cell(row=r, column=3, value=row.get("username", ""))
        ws.cell(row=r, column=4, value=row.get("first_name", ""))
        ws.cell(row=r, column=5, value=row.get("payment_type", ""))
        ws.cell(row=r, column=6, value=row.get("amount_rub", 0))
        ws.cell(row=r, column=7, value=row.get("status", ""))
        ws.cell(row=r, column=8, value=row.get("tinkoff_payment_id", ""))
        ws.cell(row=r, column=9, value=_fmt_dt(row.get("created_at")))
    _autofit(ws)
    path = _tmp_path("admin_payments_")
    wb.save(path)
    return path


# ── Отчёт 4: Использование ──

def build_usage_report(rows: list[dict]) -> Path:
    """2 листа: По пользователям (агрегированно) / Все записи."""
    wb = Workbook()

    # ── Лист 1: По пользователям ──
    ws1 = wb.active
    ws1.title = "По пользователям"
    _apply_headers(ws1, [
        "user_id", "Username", "Транскрибаций", "Всего мин",
        "Распознавание", "Тезисы", "Протокол", "Свой вариант", "Потрачено, ₽",
    ])

    # Агрегируем по пользователям
    users_agg: dict[int, dict] = {}
    for row in rows:
        uid = row.get("user_id")
        if uid not in users_agg:
            users_agg[uid] = {
                "username": row.get("username", ""),
                "count": 0, "total_min": 0.0,
                "transcribe": 0, "theses": 0, "protocol": 0, "custom": 0,
                "spent": 0,
            }
        a = users_agg[uid]
        a["count"] += 1
        a["total_min"] += float(row.get("duration_min") or 0)
        mode = row.get("mode", "transcribe")
        if mode in a:
            a[mode] += 1
        a["spent"] += int(row.get("stars_spent") or 0)

    for r, (uid, a) in enumerate(users_agg.items(), 2):
        ws1.cell(row=r, column=1, value=uid)
        ws1.cell(row=r, column=2, value=a["username"])
        ws1.cell(row=r, column=3, value=a["count"])
        ws1.cell(row=r, column=4, value=round(a["total_min"], 1))
        ws1.cell(row=r, column=5, value=a["transcribe"])
        ws1.cell(row=r, column=6, value=a["theses"])
        ws1.cell(row=r, column=7, value=a["protocol"])
        ws1.cell(row=r, column=8, value=a["custom"])
        ws1.cell(row=r, column=9, value=a["spent"])
    _autofit(ws1)

    # ── Лист 2: Все записи ──
    ws2 = wb.create_sheet("Все записи")
    _apply_headers(ws2, ["id", "user_id", "Username", "Файл", "Режим", "Мин", "Стоимость, ₽", "Пробный", "Дата"])
    for r, row in enumerate(rows, 2):
        mode_code = row.get("mode", "transcribe")
        ws2.cell(row=r, column=1, value=row.get("id", ""))
        ws2.cell(row=r, column=2, value=row.get("user_id", ""))
        ws2.cell(row=r, column=3, value=row.get("username", ""))
        ws2.cell(row=r, column=4, value=row.get("file_name", ""))
        ws2.cell(row=r, column=5, value=MODE_LABELS.get(mode_code, mode_code))
        ws2.cell(row=r, column=6, value=round(float(row.get("duration_min") or 0), 1))
        ws2.cell(row=r, column=7, value=row.get("stars_spent", 0))
        ws2.cell(row=r, column=8, value="Да" if row.get("is_trial") else "Нет")
        ws2.cell(row=r, column=9, value=_fmt_dt(row.get("created_at")))
    _autofit(ws2)

    path = _tmp_path("admin_usage_")
    wb.save(path)
    return path
